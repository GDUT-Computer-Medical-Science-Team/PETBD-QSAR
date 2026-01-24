import traceback
import pandas as pd
import xlrd
from datetime import datetime
from rdkit import Chem
from rdkit.Chem import Draw
from tqdm import tqdm
from utils import ymlReader
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
import os
import logging

log = logging.getLogger("DataMerger")
console_handler = logging.StreamHandler()
console_handler.setLevel('INFO')
fmt = u'%(asctime)s - %(funcName)s(): %(lineno)s [%(levelname)s]: %(message)s'
formatter = logging.Formatter(fmt)
console_handler.setFormatter(formatter)

log.setLevel('INFO')
log.addHandler(console_handler)


def xls2xlsx(xls_path: str) -> str:
    """
    """
    book = xlrd.open_workbook(xls_path)
    index = 0
    nrows, ncols = 0, 0
    sheet = book.sheet_by_index(0)
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    book_new = Workbook()
    default_sheet = book_new["Sheet"]
    if default_sheet is not None:
        book_new.remove(default_sheet)
    compound_index = os.path.split(os.path.splitext(xls_path)[0])[-1]
    sheet_new = book_new.create_sheet(compound_index, 0)
    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet_new.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)
    xlsx_path = os.path.splitext(xls_path)[0] + ".xlsx"
    book_new.save(xlsx_path)
    return xlsx_path


def change_suffix(filepath: str, dst_suffix: str):
    """
    :param filepath:
    :param dst_suffix:
    """
    if dst_suffix is not None:
        dot = str.find(dst_suffix, ".")
        if dot == -1:
            dst_suffix = "." + dst_suffix
        if filepath is not None:
            return os.path.splitext(filepath)[0] + dst_suffix
    return None

class DataMerger:
    def __init__(self, constants_yml_filename):
        """
        """
        self.__ymlfilename = constants_yml_filename
        self.__compound_name2mol_map = dict()
        self.__compound_name2img_map = dict()
        self.__deprecated_organ_names = ymlReader.get_deprecated_organ_names(self.__ymlfilename)
        self.__denied_organ_names = ymlReader.get_denied_organ_names(self.__ymlfilename)
        self.__time_intervals = ymlReader.get_time_intervals(self.__ymlfilename)
        self.__denied_interval_markers = ymlReader.get_denied_intervals(self.__ymlfilename)
        self.__organ_lists = ymlReader.get_target_organ_names(self.__ymlfilename)
        self.__mol_files = []

        """
        """
        cwd = os.getcwd()
        self.__raw_data_dir = os.path.join(cwd, 'data')
        if not os.path.exists(self.__raw_data_dir):
            os.makedirs(self.__raw_data_dir)
            raise FileExistsError(f"数据集目录未发现，已创建该目录：{self.__raw_data_dir}，请将数据集放入该目录后重新运行")

        cur_time = datetime.now().strftime("%Y%m%d")
        # self.__result_dir = f"{cwd}\\result\\{cur_time}"
        self.__result_dir = os.path.join(f"{cwd}", "result", f"{cur_time}")
        if not os.path.exists(self.__result_dir):
            os.makedirs(self.__result_dir)

        log_file = os.path.join(self.__result_dir, "DataMerger_DEBUG.log")
        file_handler = logging.FileHandler(log_file, encoding='utf8')
        file_handler.setLevel('DEBUG')
        file_handler.setFormatter(formatter)
        log.addHandler(file_handler)

        self.saved_pic_dir = os.path.join(self.__result_dir, 'img')
        if not os.path.exists(self.saved_pic_dir):
            os.makedirs(self.saved_pic_dir)

        data_list = os.listdir(self.__raw_data_dir)
        for file in data_list:
            if file.endswith(".mol"):
                mol_file = os.path.join(self.__raw_data_dir, file)
                compound_name = os.path.splitext(file)[0]
                self.__compound_name2mol_map[compound_name] = mol_file
                self.__mol_files.append(mol_file)

        self.output_excel_filepath = f"{self.__result_dir}\\数据表汇总.xlsx"
        if not os.path.exists(self.output_excel_filepath):
            wkc = Workbook(self.output_excel_filepath)
            wkc.save(self.output_excel_filepath)

        self.errorfile = set()

    def start_merging(self):
        """
        """
        self.__get_imgs()
        main_df = self.__init_workbook_dataframe()
        for compound_name, compound_file in tqdm(self.__compound_name2mol_map.items(), desc="正在遍历化合物数据"):
            if compound_name is not None:
                xlsx_filepath = change_suffix(compound_file, "xlsx")
                if xlsx_filepath is not None and not os.path.exists(xlsx_filepath):
                    xls_filepath = change_suffix(compound_file, "xls")
                    if not os.path.exists(xls_filepath):
                        log.error(f"化合物编号{compound_name}没有xls或xlsx数据表文件")
                        self.__save_error_compound(compound_name)
                        continue
                    log.info(f"化合物编号{compound_name}数据表格式为xls，另存为xlsx格式")
                    xlsx_filepath = xls2xlsx(xls_filepath)
                df = self.__get_DataFrame_from_workbook(xlsx_filepath)
                if df is not None:
                    try:
                        main_df = pd.concat([main_df, df], axis=0)
                    except pd.errors.InvalidIndexError as IIE:
                        log.debug(f"整合数据文件存在索引问题，对应化合物编号为{compound_name}")
                        log.debug(traceback.format_exc())
                        self.__save_error_compound(compound_file)
                        continue
                    except Exception as e:
                        log.error(f"整合数据文件出错，出错的化合物编号为{compound_name}")
                        log.error(traceback.format_exc())
                        self.__save_error_compound(compound_file)
        main_df = pd.DataFrame.dropna(main_df, axis=1, how='all')
        main_df.insert(loc=1, column='Compound structure', value="")
        main_df.insert(loc=1, column='SMILES', value="")
        main_df.to_excel(self.output_excel_filepath, index=False, engine='openpyxl', encoding='utf-8')
        log.info(f"完成化合物数据遍历，数据表保存至{self.output_excel_filepath}")

    def insert_SMILES_imgs(self):
        """
        """
        log.info("正在进行化合物结构图及SMILES插入工作，请勿打开数据表直到工作完成")
        wbc = openpyxl.load_workbook(self.output_excel_filepath)
        wsc = wbc.active

        wsc.column_dimensions['A'].width = 25
        wsc.column_dimensions['B'].width = 50
        wsc.row_dimensions[1].height = 30
        alignment = Alignment(horizontal='left', vertical='center')
        for col in wsc.columns:
            for cell in col:
                cell.alignment = alignment

        """
        """
        row = 2
        SMILES_column = 2
        for compound_name_cell in tqdm(wsc['A'], desc="正在插入化合物SMILES: "):
            compound_file_name = self.__compound_name2mol_map.get(compound_name_cell.value)
            if compound_file_name is not None:
                try:
                    writer = Chem.MolFromMolFile(compound_file_name)
                    SMILES = Chem.MolToSmiles(writer)
                except OSError as ose:
                    log.debug(f"输入的mol文件存在问题，化合物编号为{compound_file_name}")
                    log.debug(traceback.format_exc())
                    self.__save_error_compound(compound_file_name)
                    row = row + 1
                    continue
                except Exception as e:
                    log.error(f"SMILES插入出错，化合物编号为{compound_file_name}")
                    log.error(traceback.format_exc())
                    self.__save_error_compound(compound_file_name)
                    row = row + 1
                    continue
                wsc.cell(row, SMILES_column).value = SMILES
                wsc.cell(row, SMILES_column).alignment = alignment
                row = row + 1

        """
        """
        row = 2
        count = 0
        # map_length = len(self.compound_name2img_map)
        wsc.column_dimensions['C'].width = 20

        try:
            for compound_name_cell in tqdm(wsc['A'], desc="正在插入化合物结构图: "):
                # if count == map_length:
                #     break
                compound_name = compound_name_cell.value
                if compound_name == '文献编号' or compound_name == '化合物编号':
                    continue
                img_path = self.__compound_name2img_map.get(compound_name)
                if img_path is not None:
                    img = Image(img_path)
                    # img = PImage.open(img_path).resize((120, 120))

                    wsc.add_image(img, 'C' + str(row))
                    wsc.row_dimensions[row].height = 96
                    row = row + 1
                    count = count + 1
        except Exception as e:
            log.error(traceback.format_exc())
        finally:
            wbc.save(self.output_excel_filepath)
            log.info("插入工作完成，数据表保存成功")
        log.debug(f"存在问题的化合物编号: {self.errorfile}")

    def __get_imgs(self, size=(120, 120)):
        """
        """
        for mol_file in tqdm(self.__mol_files, desc="正在获取化合物结构图: "):
            split_path = os.path.splitext(mol_file)
            if split_path[-1] == '.mol':
                compound_name = os.path.split(split_path[0])[-1]
                try:
                    mol = Chem.MolFromMolFile(mol_file)
                    img_path = os.path.join(self.saved_pic_dir, compound_name + '.png')
                    # Draw.MolToImage(mol, size=(120, 120), kekulize=True)
                    Draw.MolToFile(mol, img_path, size=size)
                    self.__compound_name2img_map[compound_name] = img_path
                except OSError as ose:
                    log.debug(f"输入的mol文件存在问题，化合物编号为{compound_name}")
                    log.debug(traceback.format_exc())
                    self.__save_error_compound(compound_name)
                except Warning as e:
                    log.debug(f"生成化合物编号{compound_name}的结构图时产生警告: {traceback.format_exc()}")
                    self.__save_error_compound(compound_name)
                except Exception as e:
                    log.error(f"化合物编号{compound_name}的结构图生成出现问题:")
                    log.error(traceback.format_exc())
                    self.__save_error_compound(mol_file)
        log.info(f"化合物结构图处理完成，保存至目录: {self.saved_pic_dir}")

    def __init_workbook_dataframe(self):
        """

            Return:
        """
        log.info("初始化Dataframe表")
        headers = ['Compound index']
        for organ in self.__organ_lists:
            for time in self.__time_intervals:
                headers.append(organ + " mean" + str(time) + "min")
                headers.append(organ + " sd" + str(time) + "min")
        df = pd.DataFrame(columns=headers)
        return df

    def __get_DataFrame_from_workbook(self, workbook):
        """









            |       |30min  |60min  |

            |brain  |1      |2      |

            |blood  |0.1    |0.3    |

            ↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓↓

            |compound_index |brain 30min|brain 60min|blood 30min|blood 60min|


            Args:
            Return:
        """
        try:
            wb = openpyxl.load_workbook(workbook)
        except FileNotFoundError as e:
            log.error(traceback.format_exc())
            return None
        worksheet = wb.active
        compound_index = os.path.splitext(os.path.split(workbook)[-1])[0]
        organ_concentration_dict = dict()
        is_header_row = True
        time_headers = []

        for row in worksheet.rows:
            if is_header_row:
                for cell in row:
                    if cell.value is None:
                        continue
                    time_header = str(cell.value).strip().replace(
                        " ", "").replace("\n", "").lower()
                    if time_header in self.__denied_interval_markers:
                        continue
                    error_text = ymlReader.get_OCR_error_text(self.__ymlfilename)
                    if error_text is not None and len(error_text) > 0:
                        for k, v in error_text.items():
                            time_header = time_header.replace(k, v)
                    if not time_header.endswith("min") and not time_header.endswith("h"):
                        time_header = time_header + "min"
                    if time_header[-1] == 'h':
                        try:
                            index = time_header.find('mean')
                            if index != -1:
                                index = index + 4
                            else:
                                index = time_header.find('sd')
                                if index != -1:
                                    index = index + 2
                            if index != -1:
                                hour = int(time_header[index:-1])
                                time_header = time_header[:index] + str(hour * 60) + 'min'
                            else:
                                log.error(
                                    f"时间点数据存在缺失，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                                self.__save_error_compound(compound_index)
                                continue
                        except ValueError as e:
                            log.error(f"转换时间点数据出错，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                            log.error(traceback.format_exc())
                            self.__save_error_compound(compound_index)
                    if time_header != 'sdmin' and time_header != 'meanmin':
                        time_headers.append(time_header)
                    else:
                        log.error(f"时间点数据存在缺失，对应的化合物为{compound_index}，出错的时间点为{time_header}")
                        self.__save_error_compound(compound_index)
                # END: for cell in row:
                if len(time_headers) > 0:
                    is_header_row = False
                    if str(time_headers[0]).find('mean') == -1 and str(time_headers[0]).find('sd') == -1:
                        log.error(f"错误的列表头，对应的化合物为{compound_index}，列表头数据为{time_headers}")
                        self.__save_error_compound(compound_index)
            # END: if is_header_row:
            else:
                temp_list = []
                for cell in row:
                    if cell.value is not None:
                        temp_list.append(str(cell.value).strip()
                                         .replace("_x0001_", "")
                                         .replace(" ", "")
                                         .replace("\n", ""))
                if len(temp_list) > 0:
                    organ_name = str(temp_list[0]).lower()
                    if organ_name is not None and len(organ_name) > 0:
                        if organ_name.isalpha() and organ_name not in self.__denied_organ_names:
                            if self.__deprecated_organ_names.get(organ_name) is not None:
                                organ_name = self.__deprecated_organ_names.get(organ_name)
                            organ_concentration_dict[organ_name] = temp_list[1:]
        # END: for row in worksheet.rows

        if is_header_row is True or len(organ_concentration_dict) == 0:
            self.__save_error_compound(compound_index)
            raise ValueError(f"化合物 {compound_index} 数据存在问题")
        organs = list(organ_concentration_dict.keys())
        extended_headers = ['Compound index']
        for organ in organs:
            for time_header in time_headers:
                try:
                    extended_headers.append(str.lower(" ".join([str(organ), str(time_header)])))
                except Exception as e:
                    log.error(traceback.format_exc())
                    log.error(f"出错的化合物: {compound_index},器官名: {organ}, "
                              f"时间点: {time_header}, 当前替换后的列表头: {extended_headers}")
                    self.__save_error_compound(compound_index)
        df = pd.DataFrame(columns=extended_headers)
        df[extended_headers[0]] = [compound_index]

        for organ_name, organ_data in organ_concentration_dict.items():
            cur = 0
            for data in organ_data:
                try:
                    time_header = str.lower(' '.join([str(organ_name), str(time_headers[cur])]))
                    df[time_header] = [data]
                    cur = cur + 1
                except Exception as e:
                    self.__save_error_compound(compound_index)
                    log.error(f"Sheet rawdata: {organ_concentration_dict}")
                    log.error(f"Organs list: {organs}")
                    log.error(f"Headers list: {time_headers}")
                    log.error(f"Problem organ name: {organ_name}")
                    log.error(f"Problem organ rawdata: {data}")
                    log.error(f"Cursor index: {cur}")
                    log.error(f"Problem compound index: {compound_index}")
                    log.error(traceback.format_exc())
                    break
        return df

    def __save_error_compound(self, compound_index_or_filename: str):
        """
        :param compound_index_or_filename:
        :return:
        """
        if compound_index_or_filename is None:
            return
        split_path = os.path.splitext(compound_index_or_filename)
        if split_path[-1] == '.mol':
            compound_name = os.path.split(split_path[0])[-1]
            self.errorfile.add(compound_name)
        else:
            self.errorfile.add(compound_index_or_filename)